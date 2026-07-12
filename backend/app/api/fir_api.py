from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
import uuid
import random

from app.api.deps import get_db
from app.models.case import CaseMaster
from app.models.people import ComplainantDetails, Victim, Accused
from pydantic import BaseModel
from typing import Optional, List
from deep_translator import GoogleTranslator
import pandas as pd
import requests
from io import BytesIO
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse

router = APIRouter()

class FIRPayload(BaseModel):
    # Case Details
    district: Optional[str] = None
    police_station: Optional[str] = None
    crime_no: str
    incident_date: str
    case_category: Optional[str] = None
    latitude: float
    longitude: float
    brief_facts: str
    
    # Complainant Info
    complainant_name: str
    complainant_age: Optional[int] = None
    complainant_gender: str
    
    # Victim Info
    victim_name: Optional[str] = None
    victim_age: Optional[int] = None
    victim_gender: Optional[str] = None
    
    # Accused Info
    accused_name: Optional[str] = None
    accused_age: Optional[int] = None
    accused_gender: Optional[str] = None

from app.api.dashboard_api import STATS_CACHE
from app.models.case import CaseMaster, CrimeHead
from app.models.entities import Unit

@router.post("/add")
def add_fir(payload: FIRPayload, db: Session = Depends(get_db)):
    try:
        # Check if Crime No already exists
        existing = db.query(CaseMaster).filter(CaseMaster.CrimeNo == payload.crime_no).first()
        if existing:
            raise HTTPException(status_code=400, detail="This Crime No. already exists in the dataset.")
            
        # Generate some IDs
        case_id = db.query(func.max(CaseMaster.CaseMasterID)).scalar() or 0
        case_id += 1
        
        # Lookup CrimeHeadID
        crime_head = db.query(CrimeHead).filter(CrimeHead.CrimeGroupName == payload.case_category).first()
        crime_head_id = crime_head.CrimeHeadID if crime_head else 1
        
        # Lookup PoliceStationID
        police_station = db.query(Unit).filter(Unit.UnitName == payload.police_station).first() if payload.police_station else None
        police_station_id = police_station.UnitID if police_station else 1
        
        # 1. Insert CaseMaster
        new_case = CaseMaster(
            CaseMasterID=case_id,
            CrimeNo=payload.crime_no,
            CaseNo=f"CASE-{case_id}",
            CrimeRegisteredDate=datetime.now().date(),
            InfoReceivedPSDate=datetime.now(),
            IncidentFromDate=datetime.strptime(payload.incident_date, "%Y-%m-%d"),
            latitude=payload.latitude,
            longitude=payload.longitude,
            BriefFacts=payload.brief_facts,
            PoliceStationID=police_station_id,
            CaseCategoryID=1 if payload.case_category == "Murder" else 2,
            CrimeMajorHeadID=crime_head_id,
            CaseStatusID=1 # Under Investigation (Pending)
        )
        db.add(new_case)
        db.commit()
        
        # 2. Insert Complainant
        comp_id = db.query(func.max(ComplainantDetails.ComplainantID)).scalar() or 0
        comp_id += 1
        new_comp = ComplainantDetails(
            ComplainantID=comp_id,
            CaseMasterID=case_id,
            ComplainantName=payload.complainant_name,
            AgeYear=payload.complainant_age,
            GenderID=1 if payload.complainant_gender == 'Male' else (2 if payload.complainant_gender == 'Female' else 3)
        )
        db.add(new_comp)
        
        # 3. Insert Victim if present
        if payload.victim_name:
            vic_id = db.query(func.max(Victim.VictimMasterID)).scalar() or 0
            vic_id += 1
            new_vic = Victim(
                VictimMasterID=vic_id,
                CaseMasterID=case_id,
                VictimName=payload.victim_name,
                AgeYear=payload.victim_age,
                GenderID=1 if payload.victim_gender == 'Male' else (2 if payload.victim_gender == 'Female' else 3)
            )
            db.add(new_vic)
            
        # 4. Insert Accused if present
        if payload.accused_name:
            acc_id = db.query(func.max(Accused.AccusedMasterID)).scalar() or 0
            acc_id += 1
            new_acc = Accused(
                AccusedMasterID=acc_id,
                CaseMasterID=case_id,
                AccusedName=payload.accused_name,
                AgeYear=payload.accused_age,
                GenderID=1 if payload.accused_gender == 'Male' else (2 if payload.accused_gender == 'Female' else 3)
            )
            db.add(new_acc)
            
        db.commit()
        STATS_CACHE.clear()
        return {"status": "success", "message": "FIR added successfully", "case_id": case_id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/details/{crime_no:path}")
def get_fir_details(crime_no: str, db: Session = Depends(get_db)):
    case = db.query(CaseMaster).filter(CaseMaster.CrimeNo == crime_no).first()
    if not case:
        raise HTTPException(status_code=404, detail="FIR not found")
        
    crime_head = db.query(CrimeHead).filter(CrimeHead.CrimeHeadID == case.CrimeMajorHeadID).first()
    station = db.query(Unit).filter(Unit.UnitID == case.PoliceStationID).first()
    
    comp = db.query(ComplainantDetails).filter(ComplainantDetails.CaseMasterID == case.CaseMasterID).first()
    vic = db.query(Victim).filter(Victim.CaseMasterID == case.CaseMasterID).first()
    acc = db.query(Accused).filter(Accused.CaseMasterID == case.CaseMasterID).first()
    
    def get_gender(gid):
        if gid == 1: return "Male"
        if gid == 2: return "Female"
        if gid == 3: return "Other"
        return "Unknown"
        
    from app.models.lookups import CaseStatusMaster
    status_obj = db.query(CaseStatusMaster).filter(CaseStatusMaster.CaseStatusID == case.CaseStatusID).first()
    status_name = status_obj.CaseStatusName if status_obj else "Unknown"
    return {
        "crime_no": case.CrimeNo,
        "incident_date": case.IncidentFromDate.strftime("%Y-%m-%d") if case.IncidentFromDate else "",
        "case_category": crime_head.CrimeGroupName if crime_head else "Unknown",
        "police_station": station.UnitName if station else "Unknown",
        "latitude": case.latitude,
        "longitude": case.longitude,
        "brief_facts": case.BriefFacts or "",
        "complainant_name": comp.ComplainantName if comp else "",
        "complainant_age": comp.AgeYear if comp else None,
        "complainant_gender": get_gender(comp.GenderID) if comp else "",
        "victim_name": vic.VictimName if vic else "",
        "victim_age": vic.AgeYear if vic else None,
        "victim_gender": get_gender(vic.GenderID) if vic else "",
        "accused_name": acc.AccusedName if acc else "",
        "accused_age": acc.AgeYear if acc else None,
        "accused_gender": get_gender(acc.GenderID) if acc else "",
        "status": status_name,
    }

class TranslatePayload(BaseModel):
    text: str

@router.post("/translate")
def translate_text(payload: TranslatePayload):
    if not payload.text or not payload.text.strip():
        return {"translated_text": payload.text}
        
    try:
        translator = GoogleTranslator(source='auto', target='en')
        translated = translator.translate(payload.text)
        return {"translated_text": translated}
    except Exception as e:
        print(f"Translation error: {e}")
        return {"translated_text": payload.text}


@router.get("/bulk-template")
def get_bulk_template():
    from openpyxl.worksheet.datavalidation import DataValidation
    df = pd.DataFrame({
        "Crime No": ["CR-2023-002"],
        "Date (YYYY-MM-DD)": ["2023-10-15"],
        "Category": ["Robbery"],
        "Police Station": ["Cubbon Park Police Station"],
        "Address": ["MG Road, Bangalore, Karnataka"],
        "Brief Facts (Kannada)": ["ಎಂ.ಜಿ ರಸ್ತೆಯಲ್ಲಿ ಸರಗಳ್ಳತನ ನಡೆದಿದೆ."],
        "Complainant Name": ["Rahul Sharma"],
        "Complainant Age": [35],
        "Complainant Gender": ["Male"],
        "Victim Name": ["Priya"],
        "Victim Age": [28],
        "Victim Gender": ["Female"],
        "Accused Name": [""],
        "Accused Age": [""],
        "Accused Gender": [""]
    })
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='FIR Data', index=False)
        workbook = writer.book
        worksheet = writer.sheets['FIR Data']
        
        categories = ["Murder", "Robbery", "Cyber Fraud", "Vehicle Theft", "Chain Snatching", "Kidnapping", "Domestic Violence", "Burglary", "Drug Case", "Economic Offence", "Rioting", "Assault", "Missing Person", "Women Harassment", "Child Crime", "Theft", "Cyber Crime"]
        stations = [
            "Cubbon Park Police Station", "Ashok Nagar Police Station", "Madiwala Police Station", 
            "Whitefield Police Station", "Electronic City Police Station", "Hebbal Police Station", 
            "Jayanagar Police Station", "Yeshwanthpur Police Station", "Banashankari Police Station", 
            "Basavanagudi Police Station", "Vidhana Soudha Police Station", "Mysuru South Police Station",
            "Mysuru North Police Station", "Belagavi Rural Police Station", "Belagavi Market Police Station"
        ]
        
        lists_sheet = workbook.create_sheet('Lists')
        lists_sheet.sheet_state = 'hidden'
        
        for idx, cat in enumerate(categories, start=1):
            lists_sheet.cell(row=idx, column=1, value=cat)
            
        for idx, st in enumerate(stations, start=1):
            lists_sheet.cell(row=idx, column=2, value=st)
            
        dv_cat = DataValidation(type="list", formula1=f"='Lists'!$A$1:$A${len(categories)}", allow_blank=True)
        dv_cat.showDropDown = False
        worksheet.add_data_validation(dv_cat)
        dv_cat.add("C2:C1000")
        
        dv_st = DataValidation(type="list", formula1=f"='Lists'!$B$1:$B${len(stations)}", allow_blank=True)
        dv_st.showDropDown = False
        worksheet.add_data_validation(dv_st)
        dv_st.add("D2:D1000")
        
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="bulk_fir_template.xlsx"'
    }
    return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def geocode_address(address: str):
    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {"q": address, "format": "json"}
        headers = {"User-Agent": "CrimeVisionApp/1.0"}
        res = requests.get(url, params=params, headers=headers)
        data = res.json()
        if data:
            return float(data[0]['lat']), float(data[0]['lon'])
    except Exception as e:
        print(f"Geocoding error for {address}: {e}")
    # Default coordinates (Bangalore)
    return 12.9716, 77.5946

@router.post("/bulk-upload")
async def bulk_upload_firs(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        translator = GoogleTranslator(source='auto', target='en')
        
        inserted_count = 0
        
        for index, row in df.iterrows():
            if pd.isna(row.get('Crime No')):
                continue
                
            crime_no = str(row['Crime No'])
            
            # Skip if crime_no already exists
            if db.query(CaseMaster).filter(CaseMaster.CrimeNo == crime_no).first():
                continue
                
            address = str(row.get('Address', 'Bangalore, Karnataka'))
            lat, lng = geocode_address(address)
            
            kannada_facts = str(row.get('Brief Facts (Kannada)', ''))
            brief_facts_en = kannada_facts
            if kannada_facts.strip() and kannada_facts.strip().lower() != 'nan':
                try:
                    brief_facts_en = translator.translate(kannada_facts)
                except Exception as e:
                    print(f"Translation error: {e}")
            
            # Same DB logic as add_fir
            case_id = db.query(func.max(CaseMaster.CaseMasterID)).scalar() or 0
            case_id += 1
            
            category = str(row.get('Category', 'Other'))
            crime_head = db.query(CrimeHead).filter(CrimeHead.CrimeGroupName == category).first()
            crime_head_id = crime_head.CrimeHeadID if crime_head else 1
            
            station_name = str(row.get('Police Station', ''))
            police_station = db.query(Unit).filter(Unit.UnitName == station_name).first() if station_name else None
            police_station_id = police_station.UnitID if police_station else 1
            
            incident_date = str(row.get('Date (YYYY-MM-DD)', datetime.now().strftime("%Y-%m-%d")))
            try:
                parsed_date = datetime.strptime(incident_date[:10], "%Y-%m-%d")
            except:
                parsed_date = datetime.now()
                
            new_case = CaseMaster(
                CaseMasterID=case_id,
                CrimeNo=crime_no,
                CaseNo=f"CASE-{case_id}",
                CrimeRegisteredDate=datetime.now().date(),
                InfoReceivedPSDate=datetime.now(),
                IncidentFromDate=parsed_date,
                latitude=lat,
                longitude=lng,
                BriefFacts=brief_facts_en,
                PoliceStationID=police_station_id,
                CaseCategoryID=1 if category == "Murder" else 2,
                CrimeMajorHeadID=crime_head_id,
                CaseStatusID=1
            )
            db.add(new_case)
            
            # Complainant
            comp_name = str(row.get('Complainant Name', 'Unknown'))
            if comp_name != 'nan':
                comp_id = db.query(func.max(ComplainantDetails.ComplainantID)).scalar() or 0
                comp_id += 1
                c_gender = str(row.get('Complainant Gender', 'Male'))
                c_age = row.get('Complainant Age')
                
                new_comp = ComplainantDetails(
                    ComplainantID=comp_id,
                    CaseMasterID=case_id,
                    ComplainantName=comp_name,
                    AgeYear=int(c_age) if not pd.isna(c_age) else None,
                    GenderID=1 if c_gender == 'Male' else (2 if c_gender == 'Female' else 3)
                )
                db.add(new_comp)
                
            # Victim
            vic_name = str(row.get('Victim Name', ''))
            if vic_name and vic_name != 'nan':
                vic_id = db.query(func.max(Victim.VictimMasterID)).scalar() or 0
                vic_id += 1
                v_gender = str(row.get('Victim Gender', 'Male'))
                v_age = row.get('Victim Age')
                
                new_vic = Victim(
                    VictimMasterID=vic_id,
                    CaseMasterID=case_id,
                    VictimName=vic_name,
                    AgeYear=int(v_age) if not pd.isna(v_age) else None,
                    GenderID=1 if v_gender == 'Male' else (2 if v_gender == 'Female' else 3)
                )
                db.add(new_vic)
                
            # Accused
            acc_name = str(row.get('Accused Name', ''))
            if acc_name and acc_name != 'nan':
                acc_id = db.query(func.max(Accused.AccusedMasterID)).scalar() or 0
                acc_id += 1
                a_gender = str(row.get('Accused Gender', 'Male'))
                a_age = row.get('Accused Age')
                
                new_acc = Accused(
                    AccusedMasterID=acc_id,
                    CaseMasterID=case_id,
                    AccusedName=acc_name,
                    AgeYear=int(a_age) if not pd.isna(a_age) else None,
                    GenderID=1 if a_gender == 'Male' else (2 if a_gender == 'Female' else 3)
                )
                db.add(new_acc)
                
            inserted_count += 1
            
        db.commit()
        STATS_CACHE.clear()
        
        return {"status": "success", "message": f"Successfully imported {inserted_count} FIRs."}
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/solve/{crime_no:path}")
def solve_fir(crime_no: str, db: Session = Depends(get_db)):
    case = db.query(CaseMaster).filter(CaseMaster.CrimeNo == crime_no).first()
    if not case:
        raise HTTPException(status_code=404, detail="FIR not found")
        
    # 'Closed' status is usually 2 or 6. Let's find it.
    from app.models.lookups import CaseStatusMaster
    closed_status = db.query(CaseStatusMaster).filter(CaseStatusMaster.CaseStatusName == 'Closed').first()
    
    if closed_status:
        case.CaseStatusID = closed_status.CaseStatusID
    else:
        # Fallback if 'Closed' not found exactly
        case.CaseStatusID = 2 
        
    try:
        db.commit()
        STATS_CACHE.clear()
        return {"status": "success", "message": "Case marked as solved."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
